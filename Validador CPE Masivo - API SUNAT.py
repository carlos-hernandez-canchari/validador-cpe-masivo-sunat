#!/usr/bin/env python
# coding: utf-8

# CONSULTA MASIVA DE COMPROBANTES ELECTRÓNICOS
# • Validez vía API SUNAT
# 
# Versión funcional al 11/12/2025
# 
# Elaborado por:
# 
# Carlos Hernández Canchari
# • Analista de Planeamiento Financiero
# y Control de Gestión
# 
# LinkedIn: https://www.linkedin.com/in/carloshernandezcanchari
# 
# GitHub: https://github.com/carlos-hernandez-canchari

# 1. Importar librerías

# In[ ]:


import pandas as pd
import os
import requests
from datetime import datetime
import time
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


# 2. Configuración inicial, rutas y estilos de Excel de salida

# In[ ]:


#Ruta de Excel

EXCEL_PATH = r"Ruta de Plantilla" #En caso de trabajar con la plantilla, colocar la ruta en esta línea.
HOJA = "Consulta"
DELAY_REINTENTO = 15

#Estilos de Excel

estilo_h = Alignment(horizontal="center", vertical="center")
estilo_h_negrita = Font(bold=True)
estilo_ij = Alignment(horizontal="center", vertical="center")
estilo_k = Alignment(horizontal="left", vertical="center")


# 3. Diccionario de respuestas SUNAT

# In[ ]:


ESTADO_CP = {"0":"NO EXISTE","1":"ACEPTADO","2":"ANULADO","3":"AUTORIZADO","4":"NO AUTORIZADO"}
ESTADO_RUC = {"00":"ACTIVO","01":"BAJA PROVISIONAL","02":"BAJA PROV. POR OFICIO","03":"SUSPENSION TEMPORAL",
              "10":"BAJA DEFINITIVA","11":"BAJA DE OFICIO","22":"INHABILITADO-VENT.UNICA"}
COND_DOMI = {"00":"HABIDO","09":"PENDIENTE","11":"POR VERIFICAR","12":"NO HABIDO","20":"NO HALLADO"}


# 4. Lectura de credenciales desde plantilla de Excel + obtención del token

# In[ ]:


print("Obteniendo token de SUNAT...")
df = pd.read_excel(EXCEL_PATH, sheet_name=HOJA, header=None)

# Lectura segura (convierte NaN y celdas vacías en cadena vacía)
ruc_consultante = "" if pd.isna(df.iloc[2, 2]) else str(df.iloc[2, 2]).strip()  # Celda C3
client_id       = "" if pd.isna(df.iloc[2, 4]) else str(df.iloc[2, 4]).strip()  # Celda E3
client_secret   = "" if pd.isna(df.iloc[2, 8]) else str(df.iloc[2, 8]).strip()  # Celda I3

# === VALIDACIÓN DE CREDENCIALES ===
if not ruc_consultante:
    print("Verificar los datos de RUC en la celda C3.")
    raise SystemExit()

if not client_id:
    print("Verificar los datos de Client ID en la celda E3.")
    raise SystemExit()

if not client_secret:
    print("Verificar los datos de Client Secret en la celda I3.")
    raise SystemExit()

# =========================================================

try:
    resp_token = requests.post(
        f"https://api-seguridad.sunat.gob.pe/v1/clientesextranet/{client_id}/oauth2/token/",
        data={
            "grant_type": "client_credentials",
            "scope": "https://api.sunat.gob.pe/v1/contribuyente/contribuyentes",
            "client_id": client_id,
            "client_secret": client_secret
        },
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        timeout=30
    )
    resp_token.raise_for_status()
    token = resp_token.json()["access_token"]
    print(f"Token obtenido correctamente: {token}\n")
except Exception as e:
    print(f"Error al obtener token: {e}")
    exit()


# 5. Validaciones, consulta masiva (bucle principal de procesamiento) e inyección de resultados en plantilla

# In[ ]:


# ========================
# FUNCIÓN: VALIDAR FECHA
# ========================
def es_fecha_valida(texto):
    texto = str(texto).strip()
    if not re.match(r'^\d{1,2}/\d{1,2}/\d{4}$', texto):
        return False
    try:
        datetime.strptime(texto, "%d/%m/%Y")
        return True
    except:
        return False

while True:
    debe_repetir = False

    wb = load_workbook(EXCEL_PATH)
    ws = wb[HOJA]

    # Encontrar la última fila con datos en columnas B:G
    ultima_fila = 6
    for row_num in range(7, ws.max_row + 10):  # Comienza desde la fila 7
        celdas = [ws.cell(row=row_num, column=col).value for col in range(2, 8)]
        if any(c is not None and str(c).strip() != "" for c in celdas):
            ultima_fila = row_num

    if ultima_fila < 7:
        print("No hay datos para procesar.")
        wb.close()
        break

    fila = 7
    print("Iniciando validación de comprobantes...\n")

    while fila <= ultima_fila:
        # Leer columnas B a G desde la fila actual (7 en adelante)
        celdas = [
            ws.cell(row=fila, column=2).value,  # B
            ws.cell(row=fila, column=3).value,  # C
            ws.cell(row=fila, column=4).value,  # D
            ws.cell(row=fila, column=5).value,  # E
            ws.cell(row=fila, column=6).value,  # F
            ws.cell(row=fila, column=7).value   # G
        ]

        # Convertir None → "" y limpiar espacios
        valores = ["" if v is None else str(v).strip() for v in celdas]

        print(f"Fila {fila} → ", end="")

        # SI TODAS LAS CELDAS B:G ESTÁN VACÍAS → SALTAR Y LIMPIAR H:K
        if all(v == "" for v in valores):
            ws[f"H{fila}"].value = None
            ws[f"I{fila}"].value = None
            ws[f"J{fila}"].value = None
            ws[f"K{fila}"].value = None
            print("Fila vacía → omitida")
            fila += 1
            continue

        # Procesar la fila con datos
        ruc   = valores[0]
        tipo  = valores[1].upper()
        serie = valores[2].upper()
        num   = celdas[3]
        fecha = celdas[4]
        monto = celdas[5]

        # Normalización tipo de comprobante RxH
        if len(tipo) == 1:
            tipo = "0" + tipo
        if tipo == "02":
            tipo = "R1"

        # Fecha
        if isinstance(fecha, (datetime, pd.Timestamp)):
            fecha_str = fecha.strftime("%d/%m/%Y")
        else:
            fecha_str = str(fecha).strip() if fecha is not None else ""

        # Validaciones locales
        h = i = j = k = None

        if not str(monto).replace(".", "").replace(",", "").replace("-", "").replace(" ", "").isdigit():
            h, i, j, k = "-", "-", "-", "Importe ingresado no válido."
        elif tipo not in ["01","03","04","07","08","R1","R7"]:
            h, i, j, k = "-", "-", "-", "Tipo de comprobante ingresado no válido."
        elif len(ruc) != 11 or not ruc.isdigit():
            h, i, j, k = "-", "-", "-", "RUC ingresado no válido."
        elif len(serie) != 4:
            h, i, j, k = "-", "-", "-", "Serie ingresada no válida."
        elif tipo == "01" and not (serie.startswith("F") or serie == "E001"):
            h, i, j, k = "-", "-", "-", "Serie ingresada no válida."
        elif num is None or not str(num).isdigit():
            h, i, j, k = "-", "-", "-", "Número de comprobante ingresado no válido."
        elif not es_fecha_valida(fecha_str):
            h, i, j, k = "-", "-", "-", "Fecha ingresada no válida."
        elif tipo == "R1" and not serie.startswith("E"):
            h, i, j, k = "-", "-", "-", "Serie ingresada no válida."
        elif tipo == "03" and not (serie.startswith("B") or serie.startswith("EB")):
            h, i, j, k = "-", "-", "-", "Serie ingresada no válida."
        else:
            # CONSULTA A LA API DE SUNAT
            payload = {
                "numRuc": ruc,
                "codComp": tipo,
                "numeroSerie": serie,
                "numero": int(num),
                "fechaEmision": fecha_str,
                "monto": round(float(monto), 2)
            }
            try:
                resp = requests.post(
                    f"https://api.sunat.gob.pe/v1/contribuyente/contribuyentes/{ruc_consultante}/validarcomprobante",
                    json=payload,
                    headers={"Authorization": f"Bearer {token}"},
                    timeout=30
                )
                if resp.status_code == 200 and resp.json().get("success", False):
                    data = resp.json()["data"]
                    h = ESTADO_CP.get(data.get("estadoCp"), "DESCONOCIDO")
                    i = ESTADO_RUC.get(data.get("estadoRuc", ""), "-")
                    j = COND_DOMI.get(data.get("condDomiRuc", ""), "-")
                    obs = " ".join([o.strip("- ").strip() for o in data.get("observaciones", []) if o.strip()]) or "-"
                    k = obs
                else:
                    h, i, j, k = "SIN RESPUESTA", "-", "-", "-"
                    debe_repetir = True
            except Exception as e:
                print(f"Error conexión: {e}")
                h, i, j, k = "SIN RESPUESTA", "-", "-", "-"
                debe_repetir = True

        # Escribir resultados en columnas H, I, J, K
        ws[f"H{fila}"].value = h
        ws[f"I{fila}"].value = i
        ws[f"J{fila}"].value = j
        ws[f"K{fila}"].value = k

        # Aplicar estilos
        ws[f"H{fila}"].alignment = estilo_h
        ws[f"H{fila}"].font = estilo_h_negrita
        ws[f"I{fila}"].alignment = estilo_ij
        ws[f"J{fila}"].alignment = estilo_ij
        ws[f"K{fila}"].alignment = estilo_k

        print(f"{h}")

        fila += 1

    # Guardar archivo
    wb.save(EXCEL_PATH)
    wb.close()

    if not debe_repetir:
        print(f"\nConsultas completadas. Última fila procesada: {ultima_fila}")
        break
    else:
        print(f"\nErrores detectados → Reintentando en {DELAY_REINTENTO} segundos...")
        time.sleep(DELAY_REINTENTO)

#Aperturar archivo con las consultas resultantes.
os.startfile(EXCEL_PATH)

print("\nProceso terminado.")

